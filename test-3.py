import requests
from bs4 import BeautifulSoup
from collections import defaultdict
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# URL untuk data gempa bumi real-time
url = 'https://www.bmkg.go.id/gempabumi/gempabumi-realtime.bmkg'

# Mengirim permintaan HTTP GET ke URL
response = requests.get(url)

# Membuat objek BeautifulSoup dari konten HTML
soup = BeautifulSoup(response.content, 'html.parser')

# Mencari tabel gempa bumi real-time berdasarkan kelas CSS
table = soup.find('table', class_='table')

# Mencari semua baris dalam tabel
rows = table.find_all('tr')

# Mengumpulkan data dari setiap baris ke dalam daftar
data = []
for row in rows:
    cols = row.find_all('td')
    cols = [col.text.strip() for col in cols]
    if cols:
        data.append(cols)

# Judul kolom
headers = ['No', 'Waktu Gempa (UTC)', 'Lintang', 'Bujur', 'Magnitudo', 'Kedalaman (Km)', 'Wilayah']

# Menampilkan judul kolom
print("\n=== Data Realtime Kejadian Gempa Bumi Indonesia ===")
print(f"Waktu Data Diambil: {datetime.datetime.now()}")
print('\t'.join(headers))

# Menampilkan data yang dikumpulkan
for row in data:
    print('\t'.join(row))

# Mengelompokkan wilayah berdasarkan jumlah gempa
region_counts = defaultdict(int)
for row in data:
    if len(row) > 6:
        region = row[6]
        region_counts[region] += 1

# Menampilkan wilayah yang sering terjadi gempa
print("\n=== Wilayah yang Sering Terjadi Gempa ===")
for region, count in region_counts.items():
    print(f"Wilayah: {region}, Jumlah Gempa: {count}")

# Statistik data gempa
total_gempa = len(data)
jumlah_wilayah = len(region_counts)
rata_rata_gempa_per_wilayah = total_gempa / jumlah_wilayah

# Menentukan wilayah dengan jumlah gempa terbanyak dan terendah
most_affected_region = max(region_counts, key=region_counts.get)
least_affected_region = min(region_counts, key=region_counts.get)

# Menampilkan statistik data gempa
print("\n=== Statistik Data Gempa ===")
print(f"Total Gempa: {total_gempa}")
print(f"Jumlah Wilayah: {jumlah_wilayah}")
print(f"Rata-rata Gempa per Wilayah: {rata_rata_gempa_per_wilayah}")
print(f"Wilayah dengan Jumlah Gempa Terbanyak: {most_affected_region}, Jumlah Gempa: {region_counts[most_affected_region]}")
print(f"Wilayah dengan Jumlah Gempa Terendah: {least_affected_region}, Jumlah Gempa: {region_counts[least_affected_region]}")

# Simpan data dalam file Excel
df = pd.DataFrame(data, columns=headers)

# Buat file Excel
filename = 'data_gempa.xlsx'
df.to_excel(filename, index=False)

# Buka file Excel
workbook = Workbook()
sheet = workbook.active

# Baca data dari file Excel yang sudah dibuat
df_excel = pd.read_excel(filename)

# Tulis data ke file Excel
for r in dataframe_to_rows(df_excel, index=False, header=True):
    sheet.append(r)

# Buat grafik bar
chart = BarChart()
values = Reference(sheet, min_col=2, min_row=2, max_col=2, max_row=sheet.max_row)
categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
chart.add_data(values, titles_from_data=True)
chart.set_categories(categories)

# Letakkan grafik di sebelah kanan data
sheet.add_chart(chart, "H2")

# Simpan file Excel
workbook.save(filename)

print(f"\nData berhasil disimpan dalam file Excel: {filename}")
